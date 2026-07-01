from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponse
from django.db.models import Q, Sum
from django.core.mail import EmailMessage
from django.conf import settings
from django.template.loader import render_to_string
from django.core.paginator import Paginator
from decimal import Decimal

from .models import ClaimForm, ClaimEntry, PaymentRecord
from .forms import ClaimFormForm, ClaimEntryFormSet, ApprovalActionForm, PaymentRecordForm, SuggestedAmountForm


# ─── Role helpers ──────────────────────────────────────────────────────────────

def _is_elevated(user):
    """True for superadmin, managers, and directors — can view/delete all claims."""
    if user.is_superuser:
        return True
    role = getattr(user, 'role', '').lower() if hasattr(user, 'role') else ''
    if role in ('manager', 'director'):
        return True
    return user.groups.filter(name__in=['Manager', 'Director', 'Superadmin']).exists()


def _is_finance(user):
    """True if the user is a Finance reviewer / elevated user who can record disbursements."""
    if user.is_superuser:
        return True
    role = getattr(user, 'role', '').lower() if hasattr(user, 'role') else ''
    if role in ('finance', 'director', 'manager'):
        return True
    return user.groups.filter(name__in=['Finance', 'Director', 'Superadmin']).exists()


def _can_delete(user):
    """Only superadmin, manager, and director may delete claims."""
    return _is_elevated(user)


# ─── Email helpers ─────────────────────────────────────────────────────────────

def _send_claim_email(subject, template, context, recipient_email):
    """Send an HTML claim notification email."""
    try:
        html_body = render_to_string(template, context)
        msg = EmailMessage(
            subject=subject,
            body=html_body,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[recipient_email],
        )
        msg.content_subtype = 'html'
        msg.send(fail_silently=True)
    except Exception:
        pass


def _build_claim_url(request, claim):
    """Absolute URL to the claim detail page."""
    try:
        from django.urls import reverse
        return request.build_absolute_uri(reverse('claim_detail', kwargs={'pk': claim.pk}))
    except Exception:
        return ''


def _notify_submitted(claim, request=None):
    """
    Notify the assigned manager when an employee submits a claim.
    """
    if claim.manager and claim.manager.email:
        claim_url = _build_claim_url(request, claim) if request else ''
        _send_claim_email(
            subject=f"[BRITS] Claim submitted — {claim.employee.get_full_name()} ({claim.month.strftime('%B %Y')})",
            template='email/claim_submitted_email.html',
            context={
                'claim': claim,
                'approver_name': claim.manager.get_full_name() or claim.manager.username,
                'role': 'Manager',
                'claim_url': claim_url,
            },
            recipient_email=claim.manager.email,
        )


def _notify_manager_actioned(claim, request=None):
    """
    After manager approves → notify Finance reviewer.
    After manager rejects  → notify Employee.
    """
    claim_url = _build_claim_url(request, claim) if request else ''

    if claim.status == ClaimForm.STATUS_MANAGER_APPROVED and claim.finance_reviewer and claim.finance_reviewer.email:
        _send_claim_email(
            subject=f"[BRITS] Claim awaiting Finance approval — {claim.employee.get_full_name()} ({claim.month.strftime('%B %Y')})",
            template='email/claim_manager_actioned_email.html',
            context={
                'claim': claim,
                'recipient_name': claim.finance_reviewer.get_full_name() or claim.finance_reviewer.username,
                'claim_url': claim_url,
            },
            recipient_email=claim.finance_reviewer.email,
        )

    elif claim.status == ClaimForm.STATUS_REJECTED and claim.employee and claim.employee.email:
        _send_claim_email(
            subject=f"[BRITS] Your claim has been rejected — {claim.month.strftime('%B %Y')}",
            template='email/claim_manager_actioned_email.html',
            context={
                'claim': claim,
                'recipient_name': claim.employee.get_full_name() or claim.employee.username,
                'claim_url': claim_url,
            },
            recipient_email=claim.employee.email,
        )


def _notify_finance_outcome(claim, request=None):
    """
    After Finance approves or rejects → notify the Employee.
    """
    if claim.employee and claim.employee.email:
        claim_url = _build_claim_url(request, claim) if request else ''
        status_label = 'approved' if claim.status == ClaimForm.STATUS_FINANCE_APPROVED else 'rejected'
        _send_claim_email(
            subject=f"[BRITS] Your claim has been {status_label} — {claim.month.strftime('%B %Y')}",
            template='email/claim_finance_outcome_email.html',
            context={
                'claim': claim,
                'claim_url': claim_url,
            },
            recipient_email=claim.employee.email,
        )


# ── Shims — keep existing call sites working ──────────────────────────────────

def _notify_next_approver(claim, request=None):
    """
    Called after submit (→ manager) or after manager approval (→ finance).
    Routes to the correct helper based on claim status.
    """
    if claim.status == ClaimForm.STATUS_SUBMITTED:
        _notify_submitted(claim, request)
    elif claim.status == ClaimForm.STATUS_MANAGER_APPROVED:
        _notify_manager_actioned(claim, request)


def _notify_employee_outcome(claim, request=None):
    """
    Called after finance approves or after any rejection.
    - Finance approved or finance rejected → _notify_finance_outcome (employee gets finance email)
    - Manager rejected → _notify_manager_actioned (employee gets manager-actioned email)
    Distinguish manager vs finance rejection by checking finance_actioned_at.
    """
    if claim.status == ClaimForm.STATUS_FINANCE_APPROVED:
        _notify_finance_outcome(claim, request)
    elif claim.status == ClaimForm.STATUS_REJECTED:
        if claim.finance_actioned_at:
            # Rejected at Finance stage
            _notify_finance_outcome(claim, request)
        else:
            # Rejected at Manager stage
            _notify_manager_actioned(claim, request)


# ─── Claim list ────────────────────────────────────────────────────────────────

@login_required
def claim_list(request):
    user = request.user
    elevated = _is_elevated(user)

    """if elevated:
        all_claims_qs = ClaimForm.objects.select_related('employee', 'manager', 'finance_reviewer')
    else:
        all_claims_qs = ClaimForm.objects.filter(
            employee=user
        ).select_related('employee', 'manager', 'finance_reviewer')"""

    if elevated:
        all_claims_qs = ClaimForm.objects.select_related(
            'employee', 'manager', 'finance_reviewer'
        ).filter(
            ~Q(status=ClaimForm.STATUS_DRAFT) | Q(employee=user)
        )
    else:
        all_claims_qs = ClaimForm.objects.filter(
            employee=user
        ).select_related('employee', 'manager', 'finance_reviewer')

    search = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '').strip()
    employee_filter = request.GET.get('emp', '').strip()

    if search:
        all_claims_qs = all_claims_qs.filter(
            Q(title__icontains=search) |
            Q(employee__first_name__icontains=search) |
            Q(employee__last_name__icontains=search)
        )
    if status_filter:
        all_claims_qs = all_claims_qs.filter(status=status_filter)
    if employee_filter and elevated:
        all_claims_qs = all_claims_qs.filter(employee__id=employee_filter)

    pending_approvals = ClaimForm.objects.filter(
        Q(manager=user, status=ClaimForm.STATUS_SUBMITTED) |
        Q(finance_reviewer=user, status=ClaimForm.STATUS_PENDING_DISBURSEMENT)
    ).select_related('employee')

    paginator = Paginator(all_claims_qs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    summary = None
    if elevated:
        summary = {
            'total': all_claims_qs.count(),
            'pending': all_claims_qs.filter(
                status__in=[
                    ClaimForm.STATUS_SUBMITTED,
                    ClaimForm.STATUS_MANAGER_APPROVED,
                    ClaimForm.STATUS_PENDING_DISBURSEMENT,
                ]
            ).count(),
            'approved': all_claims_qs.filter(status=ClaimForm.STATUS_FINANCE_APPROVED).count(),
            'rejected': all_claims_qs.filter(status=ClaimForm.STATUS_REJECTED).count(),
            'total_paid': all_claims_qs.filter(
                status=ClaimForm.STATUS_FINANCE_APPROVED
            ).aggregate(s=Sum('amount_paid'))['s'] or 0,
        }

    context = {
        'page_obj': page_obj,
        'pending_approvals': pending_approvals,
        'page_title': 'All Claims' if elevated else 'My Claims',
        'elevated': elevated,
        'can_delete': _can_delete(user),
        'summary': summary,
        'search': search,
        'status_filter': status_filter,
        'status_choices': ClaimForm.STATUS_CHOICES,
        'can_view_logs': user.has_perm('core.view_fileaccesslog'),
    }
    return render(request, 'core/claims/claim_list.html', context)


# ─── Create / Edit ─────────────────────────────────────────────────────────────
@login_required
def claim_create(request):
    user = request.user

    # Default carry-forward values
    prev_claim_ref = None
    carry_amount = Decimal('0.00')

    if request.method == 'POST':

        form = ClaimFormForm(request.POST)
        formset = ClaimEntryFormSet(request.POST)

        if form.is_valid() and formset.is_valid():

            claim = form.save(commit=False)
            claim.employee = user

            # Auto title
            claim.title = claim.get_auto_title()

            # Compute carry-forward AFTER month exists
            prev_claim_ref, carry_amount = claim.compute_carry_forward()

            if prev_claim_ref:
                claim.carry_forward = carry_amount
                claim.previous_claim = prev_claim_ref

            claim.save()

            formset.instance = claim
            formset.save()

            action = request.POST.get('action', 'save')

            if action == 'submit':
                claim.status = ClaimForm.STATUS_SUBMITTED
                claim.submitted_at = timezone.now()
                claim.save()

                _notify_next_approver(claim, request)

                messages.success(request, "Claim submitted for manager approval.")
            else:
                messages.success(request, "Claim saved as draft.")

            return redirect('claim_detail', pk=claim.pk)


    else:
        from datetime import date as _date
        today = _date.today()
        auto_month = _date(today.year, today.month, 1)
        form = ClaimFormForm(
            initial={'advance': Decimal('0.00'), 'month': auto_month}
        )
        formset = ClaimEntryFormSet()

        # Show carry-forward preview for new claims
        class _TempClaim:
            employee = user
            month = auto_month
            pk = None

            def get_latest_approved_claim(self_):
                return (
                    ClaimForm.objects.filter(
                        employee=user,
                        status=ClaimForm.STATUS_FINANCE_APPROVED,
                    )
                    .order_by('-month', '-finance_actioned_at')
                    .first()
                )

            def compute_carry_forward(self_):
                prev = self_.get_latest_approved_claim()
                if prev is None:
                    return None, Decimal('0.00')
                return prev, prev.balance_due

        tmp = _TempClaim()
        prev_claim_ref, carry_amount = tmp.compute_carry_forward()

    return render(request, 'core/claims/claim_form.html', {
        'form': form,
        'formset': formset,
        'page_title': 'New Claim',
        'is_new': True,
        'prev_claim_ref': prev_claim_ref,
        'carry_amount': carry_amount,
        'can_view_logs': user.has_perm('core.view_fileaccesslog'),
    })


@login_required
def claim_edit(request, pk):
    claim = get_object_or_404(ClaimForm, pk=pk, employee=request.user)
    editable_statuses = (
        ClaimForm.STATUS_DRAFT,
        ClaimForm.STATUS_REJECTED,
    )
    if claim.status not in editable_statuses:
        messages.error(request, "Only draft or rejected claims can be edited.")
        return redirect('claim_detail', pk=pk)

    if request.method == 'POST':
        form = ClaimFormForm(request.POST, instance=claim)
        formset = ClaimEntryFormSet(request.POST, instance=claim)
        if form.is_valid() and formset.is_valid():
            claim = form.save(commit=False)
            claim.title = claim.get_auto_title()
            claim.save()
            formset.save()
            action = request.POST.get('action', 'save')
            if action == 'submit':
                claim.status = ClaimForm.STATUS_SUBMITTED
                claim.submitted_at = timezone.now()
                claim.save()
                _notify_next_approver(claim, request)
                messages.success(request, "Claim re-submitted for approval.")
            else:
                messages.success(request, "Claim updated.")
            return redirect('claim_detail', pk=pk)

    else:
        form = ClaimFormForm(instance=claim)
        formset = ClaimEntryFormSet(instance=claim)

    prev_claim_ref = claim.previous_claim
    carry_amount = claim.carry_forward

    return render(request, 'core/claims/claim_form.html', {
        'form': form,
        'formset': formset,
        'claim': claim,
        'page_title': 'Edit Claim',
        'is_new': False,
        'prev_claim_ref': prev_claim_ref,
        'carry_amount': carry_amount,
        'can_view_logs': request.user.has_perm('core.view_fileaccesslog'),
    })


# ─── Detail ────────────────────────────────────────────────────────────────────
@login_required
def claim_detail(request, pk):
    claim = get_object_or_404(
        ClaimForm.objects.select_related('employee', 'manager', 'finance_reviewer')
                         .prefetch_related('entries', 'payment_records__recorded_by'),
        pk=pk
    )

    user = request.user
    elevated = _is_elevated(user)
    is_finance_user = _is_finance(user)

    if claim.status == ClaimForm.STATUS_DRAFT and user != claim.employee:
        messages.error(request, "You do not have access to this claim.")
        return redirect('claim_list')

    allowed = [claim.employee, claim.manager, claim.finance_reviewer]

    if user not in allowed and not elevated:
        messages.error(request, "You do not have access to this claim.")
        return redirect('claim_list')

    prev_claim = claim.get_previous_claim()
    #claim_history = claim.get_claim_history()[:5]
    claim_history_qs = claim.get_claim_history()
    history_paginator = Paginator(claim_history_qs, 5)
    history_page_obj = history_paginator.get_page(request.GET.get('history_page'))

    # ── Finance Actions — two stages ─────────────────────────────────────────────
    # Stage 1 (manager, status = manager_approved): manager records the amount
    #   they recommend Finance disburse ("Amount Suggested"). This forwards the
    #   claim to Finance (status -> pending_disbursement). No money is marked
    #   as paid and the claim is not closed.
    # Stage 2 (finance, status = pending_disbursement): finance reviews /
    #   edits the suggested amount and records the final payment. This creates
    #   the official PaymentRecord, marks the claim Closed, and redirects back
    #   to the claim list.
    suggestion_form = None
    payment_form = None
    payment_records = claim.payment_records.all()

    can_suggest_amount = (
        claim.status == ClaimForm.STATUS_MANAGER_APPROVED
        and (claim.manager == user or elevated)
    )
    can_record_payment = (
        claim.status == ClaimForm.STATUS_PENDING_DISBURSEMENT
        and (claim.finance_reviewer == user or _is_finance_strict(user))
    )

    if can_suggest_amount and request.method == 'POST' and 'record_suggestion' in request.POST:
        suggestion_form = SuggestedAmountForm(request.POST, instance=claim)
        if suggestion_form.is_valid():
            amount = suggestion_form.cleaned_data['suggested_amount']
            claim.record_suggestion(user, amount)
            messages.success(
                request,
                f"Suggested amount of KES {amount:,.2f} recorded. "
                f"Claim forwarded to Finance for payment."
            )
            return redirect('claim_detail', pk=pk)
    elif can_suggest_amount:
        suggestion_form = SuggestedAmountForm()

    if can_record_payment and request.method == 'POST' and 'record_payment' in request.POST:
        payment_form = PaymentRecordForm(request.POST)
        if payment_form.is_valid():
            record = payment_form.save(commit=False)
            record.claim = claim
            record.recorded_by = user
            record.paid_at = timezone.now()
            record.save()  # triggers claim.refresh_amount_paid() automatically

            claim.refresh_from_db()
            claim.status = ClaimForm.STATUS_FINANCE_APPROVED
            claim.finance_actioned_at = timezone.now()
            claim.save()

            _notify_employee_outcome(claim, request)

            overpayment = claim.overpayment
            if overpayment > Decimal('0.00'):
                messages.warning(
                    request,
                    f"⚠️ Payment of KES {record.amount:,.2f} recorded and the claim is now closed. "
                    f"This exceeds the net payable by KES {overpayment:,.2f}. "
                    f"The surplus will automatically be deducted from "
                    f"{claim.employee.get_full_name()}'s next claim as a carry-forward advance."
                )
            else:
                messages.success(
                    request,
                    f"Payment of KES {record.amount:,.2f} recorded. Claim closed."
                )

            return redirect('claim_list')
    elif can_record_payment:
        initial = {}
        if claim.suggested_amount:
            initial['amount'] = claim.suggested_amount
        payment_form = PaymentRecordForm(initial=initial)

    return render(request, 'core/claims/claim_detail.html', {
        'claim': claim,
        'entries': claim.entries.all(),
        'approval_form': ApprovalActionForm(),
        'can_approve': claim.can_approve(user),
        'page_title': f"Claim — {claim.employee.get_full_name()}",
        'can_view_logs': user.has_perm('core.view_fileaccesslog'),
        'elevated': elevated,
        'is_finance_user': is_finance_user,
        'can_delete': _can_delete(user),
        'prev_claim': prev_claim,
        #'claim_history': claim_history,
        'claim_history': history_page_obj,
        'history_page_obj': history_page_obj,
        'suggestion_form': suggestion_form,
        'can_suggest_amount': can_suggest_amount,
        'payment_form': payment_form,
        'payment_records': payment_records,
        'can_record_payment': can_record_payment,
        'can_view_payment_log': elevated or is_finance_user or claim.finance_reviewer == user,
    })

def _is_finance_strict(user):
    if user.is_superuser:
        return True
    role = getattr(user, 'role', '').lower() if hasattr(user, 'role') else ''
    if role in ('finance', 'director'):   # 'manager' removed
        return True
    return user.groups.filter(name__in=['Finance', 'Director', 'Superadmin']).exists()

# ─── Approval ──────────────────────────────────────────────────────────────────

@login_required
def claim_approve(request, pk):
    if request.method != 'POST':
        return redirect('claim_detail', pk=pk)
    claim = get_object_or_404(ClaimForm, pk=pk)
    if not claim.can_approve(request.user):
        messages.error(request, "You are not authorised to approve this claim at this stage.")
        return redirect('claim_detail', pk=pk)
    form = ApprovalActionForm(request.POST)
    if form.is_valid():
        claim.approve(request.user, comment=form.cleaned_data.get('comment', ''))
        _notify_next_approver(claim, request)
        # Only notify the employee when Finance is the final approver
        if claim.status == ClaimForm.STATUS_FINANCE_APPROVED:
            _notify_employee_outcome(claim, request)
        messages.success(request, f"Claim approved. Status: {claim.get_status_display()}")
    return redirect('claim_detail', pk=pk)


@login_required
def claim_reject(request, pk):
    if request.method != 'POST':
        return redirect('claim_detail', pk=pk)
    claim = get_object_or_404(ClaimForm, pk=pk)
    if not claim.can_approve(request.user):
        messages.error(request, "You are not authorised to action this claim.")
        return redirect('claim_detail', pk=pk)
    form = ApprovalActionForm(request.POST)
    if form.is_valid():
        comment = form.cleaned_data.get('comment', '')
        if not comment:
            messages.error(request, "A comment is required when rejecting a claim.")
            return redirect('claim_detail', pk=pk)
        claim.reject(request.user, comment=comment)
        _notify_employee_outcome(claim, request)
        messages.warning(request, "Claim rejected. Employee has been notified.")
    return redirect('claim_detail', pk=pk)


# ─── Delete ────────────────────────────────────────────────────────────────────

@login_required
def claim_delete(request, pk):
    if not _can_delete(request.user):
        messages.error(request, "You do not have permission to delete claims.")
        return redirect('claim_detail', pk=pk)
    claim = get_object_or_404(ClaimForm, pk=pk)
    if request.method == 'POST':
        employee_name = claim.employee.get_full_name()
        month_label = claim.month.strftime('%B %Y')
        claim.delete()
        messages.success(request, f"Claim for {employee_name} ({month_label}) has been deleted.")
        return redirect('claim_list')
    return redirect('claim_detail', pk=pk)


# ─── Export ────────────────────────────────────────────────────────────────────

@login_required
def claim_export_pdf(request, pk):
    claim = get_object_or_404(
        ClaimForm.objects.select_related('employee').prefetch_related('entries', 'payment_records'),
        pk=pk
    )
    if claim.status == ClaimForm.STATUS_DRAFT and request.user != claim.employee:
        return HttpResponse("Forbidden", status=403)
    allowed = [claim.employee, claim.manager, claim.finance_reviewer]
    if request.user not in allowed and not _is_elevated(request.user):
        return HttpResponse("Forbidden", status=403)
    try:
        from weasyprint import HTML
        html_string = render_to_string('core/claims/claim_pdf.html', {'claim': claim})
        pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
        response = HttpResponse(pdf_file, content_type='application/pdf')
        filename = f"claim_{claim.employee.get_full_name().replace(' ', '_')}_{claim.month.strftime('%Y_%m')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except ImportError:
        return HttpResponse("WeasyPrint is not installed. Run: pip install weasyprint", status=500)

@login_required
def claim_export_excel(request, pk):
    claim = get_object_or_404(
        ClaimForm.objects.select_related('employee').prefetch_related('entries', 'payment_records'),
        pk=pk
    )
    if claim.status == ClaimForm.STATUS_DRAFT and request.user != claim.employee:
        return HttpResponse("Forbidden", status=403)
    allowed = [claim.employee, claim.manager, claim.finance_reviewer]
    if request.user not in allowed and not _is_elevated(request.user):
        return HttpResponse("Forbidden", status=403)
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Claim Form"

        header_font = Font(bold=True, size=13)
        bold = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center')
        right = Alignment(horizontal='right')
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        navy_fill = PatternFill("solid", fgColor="1F3864")
        navy_font = Font(bold=True, color="FFFFFF")

        def style_cell(cell, font=None, align=None, fill=None, border_=None):
            if font: cell.font = font
            if align: cell.alignment = align
            if fill: cell.fill = fill
            if border_: cell.border = border_

        ws.merge_cells('A1:L1')
        ws['A1'] = f"{claim.display_title.upper()} — {claim.month.strftime('%B %Y').upper()}"
        style_cell(ws['A1'], font=header_font, align=center)
        ws.row_dimensions[1].height = 28

        headers = ['DATE', 'SITE', 'TICKET NO.', 'TO', 'FROM',
           'BREAKFAST', 'LUNCH', 'DINNER', 'BED', 'OTHER EXPENDITURE', 'TOTAL', 'ADVANCE']
        col_widths = [14, 22, 16, 10, 10, 12, 10, 10, 10, 16, 12, 12]
        for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=2, column=col_idx, value=h)
            style_cell(cell, font=navy_font, align=center, fill=navy_fill, border_=border)
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        ws.row_dimensions[2].height = 20

        for row_idx, entry in enumerate(claim.entries.all(), start=3):
            row_data = [
                entry.date, entry.site, entry.ticket_number,
                float(entry.transport_to), float(entry.transport_from),
                float(entry.breakfast), float(entry.lunch),
                float(entry.dinner), float(entry.bed), float(entry.other_expenditure),
                float(entry.total), '',
            ]
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border
                if col_idx == 1:
                    cell.number_format = 'DD/MM/YYYY'
                    cell.alignment = center
                elif col_idx >= 4:
                    cell.number_format = '#,##0.00'
                    cell.alignment = right

        last_data_row = 2 + claim.entries.count()
        subtotal_row = last_data_row + 2

        ws.merge_cells(f'A{subtotal_row}:J{subtotal_row}')
        ws.cell(subtotal_row, 1, 'Sub total').font = bold
        ws.cell(subtotal_row, 1).alignment = right
        ws.cell(subtotal_row, 11, float(claim.subtotal)).number_format = '#,##0.00'
        ws.cell(subtotal_row, 11).font = bold
        ws.cell(subtotal_row, 12, float(claim.advance)).number_format = '#,##0.00'
        ws.cell(subtotal_row, 12).font = bold

        if claim.carry_forward:
            cf_row = subtotal_row + 1
            ws.merge_cells(f'A{cf_row}:J{cf_row}')
            ws.cell(cf_row, 1, 'Carry-forward from previous claim').font = bold
            ws.cell(cf_row, 1).alignment = right
            ws.cell(cf_row, 11, float(claim.carry_forward)).number_format = '#,##0.00'
            ws.cell(cf_row, 11).font = bold
            tma_row = cf_row + 1
        else:
            tma_row = subtotal_row + 1

        ws.merge_cells(f'A{tma_row}:J{tma_row}')
        ws.cell(tma_row, 1, 'Total Minus Advance').font = bold
        ws.cell(tma_row, 1).alignment = right
        ws.cell(tma_row, 11, float(claim.total_minus_advance)).number_format = '#,##0.00'
        ws.cell(tma_row, 11).font = bold

        pr_start = tma_row + 1
        for i, pr in enumerate(claim.payment_records.all()):
            pr_row = pr_start + i
            ws.merge_cells(f'A{pr_row}:J{pr_row}')
            label = f"Payment #{i+1} — {pr.paid_at.strftime('%d %b %Y')}"
            if pr.note:
                label += f" ({pr.note})"
            ws.cell(pr_row, 1, label).font = bold
            ws.cell(pr_row, 1).alignment = right
            ws.cell(pr_row, 11, float(pr.amount)).number_format = '#,##0.00'
            ws.cell(pr_row, 11).font = bold

        paid_row = pr_start + claim.payment_records.count()
        ws.merge_cells(f'A{paid_row}:J{paid_row}')
        ws.cell(paid_row, 1, 'Total Amount Paid').font = bold
        ws.cell(paid_row, 1).alignment = right
        ws.cell(paid_row, 11, float(claim.amount_paid)).number_format = '#,##0.00'
        ws.cell(paid_row, 11).font = bold

        if claim.overpayment > 0:
            op_row = paid_row + 1
            ws.merge_cells(f'A{op_row}:J{op_row}')
            ws.cell(op_row, 1, 'Overpayment (carry-forward to next claim)').font = bold
            ws.cell(op_row, 1).alignment = right
            ws.cell(op_row, 11, float(claim.overpayment)).number_format = '#,##0.00'
            ws.cell(op_row, 11).font = Font(bold=True, color='FF0000')
            sig_row = op_row + 3
        elif claim.balance_due < 0:
            # Negative balance_due = employer still owes employee
            bd_row = paid_row + 1
            ws.merge_cells(f'A{bd_row}:J{bd_row}')
            ws.cell(bd_row, 1, 'Outstanding Balance').font = bold
            ws.cell(bd_row, 1).alignment = right
            ws.cell(bd_row, 11, float(abs(claim.balance_due))).number_format = '#,##0.00'
            ws.cell(bd_row, 11).font = Font(bold=True, color='FF0000')
            sig_row = bd_row + 3
        else:
            sig_row = paid_row + 3

        ws.cell(sig_row, 1, 'EMPLOYEE SIGN').font = bold
        ws.cell(sig_row, 5, 'DATE').font = bold
        ws.cell(sig_row + 1, 1, 'MANAGER SIGN').font = bold
        ws.cell(sig_row + 1, 5, 'DATE').font = bold
        ws.cell(sig_row + 2, 1, 'FINANCE SIGN').font = bold
        ws.cell(sig_row + 2, 5, 'DATE').font = bold

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"claim_{claim.employee.get_full_name().replace(' ', '_')}_{claim.month.strftime('%Y_%m')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    except ImportError:
        return HttpResponse("openpyxl is not installed. Run: pip install openpyxl", status=500)
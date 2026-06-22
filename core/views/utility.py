from .imports import *

@login_required
def profile_view(request):
    context = {
        'user': request.user,
        'user_form': UserUpdateForm(instance=request.user),
        'profile_form': ProfileUpdateForm(instance=request.user.profile),
    }

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'accounts/profile_content.html', context)

    return render(request, 'accounts/profile.html', context)


class SettingsView(View):
    def get(self, request):
        user_form = UserUpdateForm(instance=request.user)
        profile, created = Profile.objects.get_or_create(user=request.user)
        profile_form = ProfileUpdateForm(instance=profile)

        return render(request, 'accounts/settings.html', {
            'user_form': user_form,
            'profile_form': profile_form
        })

    def post(self, request):
        user_form = UserUpdateForm(request.POST, instance=request.user)
        profile, created = Profile.objects.get_or_create(user=request.user)
        profile_form = ProfileUpdateForm(request.POST, request.FILES, instance=profile)

        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            messages.success(request, "Your settings have been updated.")
            return redirect('settings') 

        print("FILES:", request.FILES)

        return render(request, 'accounts/settings.html', {
            'user_form': user_form,
            'profile_form': profile_form
        })

"""def export_tickets_to_excel(tickets):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"

    ws.append([
        "ID", "Terminal", "Issue", "Status",
        "Assigned To", "Created At"
    ])

    for ticket in tickets:
        ws.append([
            ticket.id,
            ticket.terminal.branch_name if ticket.terminal else "",
            ticket.title,
            ticket.get_status_display(),
            ticket.assigned_to.username if ticket.assigned_to else "",
            ticket.created_at.strftime("%Y-%m-%d %H:%M"),
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=tickets.xlsx"
    wb.save(response)

    return response"""

def export_tickets_to_excel(tickets, customer_name=None, terminal_name=None, start_date=None, end_date=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"

    ws.append([
        "ID", "Terminal", "Issue", "Status",
        "Assigned To", "Created At"
    ])

    for ticket in tickets:
        ws.append([
            ticket.id,
            ticket.terminal.branch_name if ticket.terminal else "",
            ticket.title,
            ticket.get_status_display(),
            ticket.assigned_to.username if ticket.assigned_to else "",
            ticket.created_at.strftime("%Y-%m-%d %H:%M"),
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=tickets.xlsx"
    wb.save(response)
    return response

def export_report_tickets_to_excel(tickets, customer_name=None, terminal_name=None, problem_category_name=None, start_date=None, end_date=None):
    import openpyxl
    from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
    from django.http import HttpResponse
    from django.utils import timezone

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Tickets'

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Define header fill (light gray)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Define headers
    headers = [
        'Customer', 'Terminal', 'Problem Category', 'Issue Type', 'Description', 'Status',
        'Assigned To', 'Resolved By', 'Resolution', 'Created At', 'Updated At',
        'Resolved At', 'Comments'
    ]
    sheet.append(headers)

    # Style headers (bold + background + border)
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Populate rows
    for ticket in tickets:
        comments_text = "\n".join([
            f"{comment.created_by.username if comment.created_by else 'Unknown'} ({comment.created_at.strftime('%Y-%m-%d')}): {comment.content}"
            for comment in ticket.comments.all()
        ])

        row_data = [
            ticket.customer.name if ticket.customer else "",
            ticket.terminal.branch_name if ticket.terminal else "",
            str(ticket.problem_category) if ticket.problem_category else "",
            ticket.title or "",
            ticket.description or "",
            ticket.status or "",
            str(ticket.assigned_to) if ticket.assigned_to else "",
            str(ticket.resolved_by) if hasattr(ticket, 'resolved_by') and ticket.resolved_by else "",
            ticket.resolution or "",
            ticket.created_at.strftime('%Y-%m-%d %H:%M') if ticket.created_at else "",
            ticket.updated_at.strftime('%Y-%m-%d %H:%M') if ticket.updated_at else "",
            ticket.resolved_at.strftime('%Y-%m-%d %H:%M') if hasattr(ticket, 'resolved_at') and ticket.resolved_at else "",
            comments_text
        ]

        sheet.append(row_data)

    # Wrap text in Comments column
    comment_col_index = headers.index('Comments') + 1
    for row in sheet.iter_rows(min_row=2, min_col=comment_col_index, max_col=comment_col_index):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = thin_border

    # Apply borders to all cells
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top')

    # Determine filename
    name_part = "report"
    if customer_name:
        name_part = f"{customer_name.replace(' ', '_')}_report"
    elif terminal_name:
        name_part = f"{terminal_name.replace(' ', '_')}_report"
    elif problem_category_name:
        name_part = f"{problem_category_name.replace(' ', '_')}_report"

    date_part = ''
    if start_date and end_date:
        date_part = f"{start_date}_to_{end_date}"
    elif start_date:
        date_part = f"from_{start_date}"
    elif end_date:
        date_part = f"to_{end_date}"

    filename = f"{name_part}_{date_part or timezone.now().strftime('%Y-%m-%d')}.xlsx"

    # Send response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    workbook.save(response)

    return response
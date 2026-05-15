from .imports import *

@login_required(login_url='login')
def pre_dashboards(request):
    # Check if the user is assigned as either overseer or custodian in any customer
    is_overseer_or_custodian = False

    # Check if the user is an overseer or custodian
    if Customer.objects.filter(overseer=request.user).exists() or Terminal.objects.filter(custodian=request.user).exists():
        is_overseer_or_custodian = True

    print(f"Is overseer or custodian: {is_overseer_or_custodian}") 

    return render(request, 'core/pre_dashboards.html', {
        'is_overseer_or_custodian': is_overseer_or_custodian
    })

@login_required
def ticketing_dashboard(request):
    now = timezone.localtime(timezone.now())
    day_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = day_start - timedelta(days=day_start.weekday())
    month_start = day_start.replace(day=1)
    year_start = day_start.replace(month=1, day=1)

    # ======================================
    # Role-Based Filtering
    # ======================================
    ticket_filter = Ticket.objects.none()
    profile = getattr(request.user, 'profile', None)

    if Customer.objects.filter(custodian=request.user).exists() or Terminal.objects.filter(custodian=request.user).exists():
        user_group = "Custodian"
    elif Customer.objects.filter(overseer=request.user).exists():
        user_group = "Overseer"
    elif request.user.groups.filter(name="Director").exists():
        user_group = "Director"
    elif request.user.groups.filter(name="Manager").exists():
        user_group = "Manager"
    elif request.user.groups.filter(name="Staff").exists():
        user_group = "Staff"
    elif request.user.is_superuser:
        user_group = "Superuser"
    else:
        user_group = "Customer"
        
    hide_region = user_group == "Custodian"
    hide_terminal = user_group == "Custodian"
    hide_customer = user_group in ["Custodian", "Overseer"]

    if request.user.is_superuser or request.user.groups.filter(name__in=['Director', 'Manager', 'Staff']).exists():
        # Superusers and internal staff see all tickets
        ticket_filter = Ticket.objects.all()
    else:
        # For overseer: filter by customer they oversee
        customer = Customer.objects.filter(overseer=request.user).first()
        if customer:
            ticket_filter = Ticket.objects.filter(customer=customer)
        elif profile and profile.terminal:
            ticket_filter = Ticket.objects.filter(terminal=profile.terminal)
            print(f"{request.user.username} is Custodian for {profile.terminal.branch_name} with customer {profile.customer.name}")
        else:
            print(f"{request.user.username} has no profile or terminal")

    # Time-based ticket counts (restricted)
    time_data = {
        'day': ticket_filter.filter(created_at__gte=day_start).count(),
        'week': ticket_filter.filter(created_at__gte=week_start).count(),
        'month': ticket_filter.filter(created_at__gte=month_start).count(),
        'year': ticket_filter.filter(created_at__gte=year_start).count(),
    }

    # Aggregated counts (status & priority)
    priority_counts = list(ticket_filter.values('priority').annotate(count=Count('id')))
    status_counts = list(ticket_filter.values('status').annotate(count=Count('id')))
    # >>> ADD SORTING HERE <<<
    PRIORITY_ORDER = ['low', 'medium', 'high', 'critical']
    STATUS_ORDER = ['open', 'in_progress', 'closed']

    priority_counts.sort(
        key=lambda x: PRIORITY_ORDER.index(x['priority'].lower())
                    if x['priority'] and x['priority'].lower() in PRIORITY_ORDER else len(PRIORITY_ORDER)
    )

    status_counts.sort(
        key=lambda x: STATUS_ORDER.index(x['status'].lower().replace(' ', '_'))
                     if x['status'] and x['status'].lower().replace(' ', '_') in STATUS_ORDER else len(STATUS_ORDER)
    )
    # >>> END OF ADDITION <<<

    # Monthly ticket trends
    monthly_trends = (
        ticket_filter
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )

    # Top terminals with most tickets
    terminal_data = (
        ticket_filter
        .values('terminal__id', 'terminal__branch_name')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )

    # Region trends
    region_data = (
        ticket_filter
        .values('terminal__region__name')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    # Categories
    category_data = (
        ticket_filter
        .values('problem_category__name')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )

    # Top customers
    customer_data = (
        ticket_filter
        .values('terminal__customer__name')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )

    # Overview + Category-time widgets
    overview_data = [
        {'label': 'Today', 'count': time_data['day']},
        {'label': 'This Week', 'count': time_data['week']},
        {'label': 'This Month', 'count': time_data['month']},
        {'label': 'This Year', 'count': time_data['year']},
    ]

    category_time_data = [
        {'category': d['problem_category__name'], 'daily_count': d['count']} 
        for d in category_data  
    ]

    kpi_data = [
        ("Today", "dailyCount", "fa-sun"),
        ("This Week", "weeklyCount", "fa-calendar-week"),
        ("This Month", "monthlyCount", "fa-calendar-alt"),
        ("This Year", "yearlyCount", "fa-calendar"),
    ]

    # User Group Determination
    user_group = None
    if Customer.objects.filter(custodian=request.user).exists():
        user_group = "Custodian"
    elif Customer.objects.filter(overseer=request.user).exists():
        user_group = "Overseer"
    else:
        if request.user.groups.filter(name="Director").exists():
            user_group = "Director"
        elif request.user.groups.filter(name="Manager").exists():
            user_group = "Manager"
        elif request.user.groups.filter(name="Staff").exists():
            user_group = "Staff"
        else:
            user_group = "Customer"  

    allowed_roles = ["Director", "Manager", "Staff", "Superuser"]

    context = {
        'user_group': user_group,
        "allowed_roles": allowed_roles,
        "kpi_data": kpi_data,
        "hide_region": hide_region,
        "hide_terminal": hide_terminal,
        "hide_customer": hide_customer,
        'status_data': json.dumps(list(status_counts)),
        'priority_data': json.dumps(list(priority_counts)),
        'monthly_data': json.dumps([
            {'month': calendar.month_abbr[d['month'].month], 'count': d['count']}
            for d in monthly_trends if d['month']
        ]),
        'terminal_data': json.dumps([
            {'terminal_id': d['terminal__id'], 'terminal': d['terminal__branch_name'], 'count': d['count']}
            for d in terminal_data
        ]),
        'region_data': json.dumps([
            {'region': d['terminal__region__name'], 'count': d['count']}
            for d in region_data
        ]),
        'time_data': json.dumps(time_data),
        'category_data': json.dumps([
            {'category': d['problem_category__name'], 'count': d['count']}
            for d in category_data
        ]),
        'customer_data': json.dumps([
            {'customer': d['terminal__customer__name'], 'count': d['count']}
            for d in customer_data
        ]),
        'overview_data': json.dumps(overview_data),
        'category_time_data': json.dumps(category_time_data),
    }

    return render(request, 'core/helpdesk/ticketing_dashboard.html', context)

@login_required(login_url='login')
def statistics_view(request):
    today = timezone.now()
    print(f"timezone.now() in view: {today} (aware: {timezone.is_aware(today)})")

    tickets = Ticket.objects.all()
    customer_id = request.GET.get("customer")
    region_id = request.GET.get("region")
    terminal_id = request.GET.get("terminal")

    user_group = None
    assigned_customer = None
    assigned_terminal = None
    assigned_region = None
    user = request.user
    user_profile = getattr(user, 'profile', None)

    
    # --- Role-based Filtering ---
    if user.is_superuser or user.groups.filter(name__in=['Director', 'Manager', 'Staff']).exists():
        user_group = "Internal"
    elif Customer.objects.filter(overseer=user).exists():
        user_group = "Overseer"
        assigned_customer = Customer.objects.filter(overseer=user).first()
        if assigned_customer:
            print(f"{user.username} is Overseer for {assigned_customer.name}")
            tickets = tickets.filter(terminal__customer=assigned_customer)
        else:
            tickets = Ticket.objects.none()
    elif user_profile and user_profile.terminal:
        if user_profile.terminal.custodian == user:
            user_group = "Custodian"
            assigned_terminal = user_profile.terminal
            assigned_customer = assigned_terminal.customer
            assigned_region = assigned_terminal.region
            print(f"{user.username} is Custodian for {assigned_terminal.branch_name}")
            tickets = tickets.filter(terminal=assigned_terminal)
        else:
            tickets = Ticket.objects.none()
    else:
        tickets = Ticket.objects.none()

    hide_sla = user_group in ["Custodian", "Overseer"]
    hide_assignee = user_group in ["Custodian", "Overseer"]
    hide_resolver = user_group in ["Custodian", "Overseer"]
    hide_terminal = user_group == "Custodian"


    # --- Filters ---
    time_period = request.GET.get('time-period', "all_time")
    customer_filter = request.GET.get('customer', 'all')
    terminal_filter = request.GET.get('terminal', 'all')
    region_filter = request.GET.get('region', 'all')

    if time_period == 'today':
        start_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = today.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif time_period == 'yesterday':
        start_date = (today - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = (today - timedelta(days=1)).replace(hour=23, minute=59, second=59, microsecond=999999)
    elif time_period == 'lastweek':
        start_date = today - timedelta(days=today.weekday() + 7)
        start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = start_date + timedelta(days=6, hours=23, minutes=59, seconds=59, microseconds=999999)
    elif time_period == 'lastmonth':
        end_date = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0) - timedelta(microseconds=1)
        start_date = end_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif time_period == 'lastyear':
        start_date = today.replace(year=today.year - 1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        end_date = today.replace(year=today.year - 1, month=12, day=31, hour=23, minute=59, second=59, microsecond=999999)
    elif time_period == 'all_time':
        start_date = None
        end_date = None
    else:
        start_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = today.replace(hour=23, minute=59, second=59, microsecond=999999)

    

    # Filter by customer
    if customer_filter not in ['all', '', None]:
        try:
            tickets = tickets.filter(terminal__customer__id=int(customer_filter))
        except ValueError:
            pass


    # Filter by terminal
    if terminal_filter not in ['all', '', None]:
        try:
            tickets = tickets.filter(terminal__id=int(terminal_filter))
        except ValueError:
            pass

    # Filter by region
    if region_filter not in ['all', '', None]:
        try:
            tickets = tickets.filter(terminal__region__id=int(region_filter))
        except ValueError:
            pass    

    # Filter by time period
    if time_period != 'all_time' and start_date and end_date:
        tickets = tickets.filter(created_at__range=[start_date, end_date])


    tickets_list = list(tickets.iterator())
    ticket_statuses = tickets.values('status').annotate(status_count=Count('status'))
    status_labels = [status['status'] for status in ticket_statuses]
    status_counts = [status['status_count'] for status in ticket_statuses]

    days = [today - timedelta(days=i) for i in range(7)]
    tickets_per_day = [tickets.filter(created_at__date=day.date()).count() for day in days]
    weekdays = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    tickets_per_weekday = [tickets.filter(created_at__week_day=(i % 7) + 1).count() for i in range(7)]
    hours = [f"{i}-{i+1}" for i in range(24)]
    tickets_per_hour = [tickets.filter(created_at__hour=i).count() for i in range(24)]
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    tickets_per_month = [tickets.filter(created_at__month=i+1).count() for i in range(12)]
    years = sorted(list(set(ticket.created_at.year for ticket in tickets_list)))
    tickets_per_year = [tickets.filter(created_at__year=year).count() for year in years]
    tickets_per_terminal = tickets.values('terminal__branch_name').annotate(ticket_count=Count('id'))
    ticket_categories = tickets.values('problem_category__name').annotate(ticket_count=Count('id'))

    if request.GET.get("export") == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Statistics"

        headers = ["Terminal", "Ticket Count"]
        ws.append(headers)

        thin_border = Loader(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for entry in tickets_per_terminal:
            ws.append([entry['terminal__branch_name'], entry['ticket_count']])

        # Apply borders to all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="statistics.xlsx"'
        wb.save(response)
        return response

    available_customers, available_terminals, available_regions = [], [], []
    if user_group == "Internal":
        available_customers = list(Customer.objects.values('id', 'name'))
        available_terminals = list(Terminal.objects.select_related('customer', 'region').values(
            'id', 'branch_name', 'customer__name', 'region__name', 'region__id'
        ))
        available_regions = list(Region.objects.values('id', 'name'))
    elif user_group == "Overseer" and assigned_customer:
        available_customers = [{'id': assigned_customer.id, 'name': assigned_customer.name}]
        available_terminals = list(Terminal.objects.filter(customer=assigned_customer).select_related('customer', 'region').values(
            'id', 'branch_name', 'customer__name', 'region__name', 'region__id'
        ))
        available_regions = list(Region.objects.filter(terminal__customer=assigned_customer).distinct().values('id', 'name'))
    elif user_group == "Custodian" and assigned_terminal:
        available_customers = [{'id': assigned_customer.id, 'name': assigned_customer.name}]
        available_terminals = [{
            'id': assigned_terminal.id,
            'branch_name': assigned_terminal.branch_name,
            'customer__name': assigned_terminal.customer.name if assigned_terminal.customer else 'N/A',
            'region__name': assigned_terminal.region.name if assigned_terminal.region else 'N/A',
            'region_id': assigned_terminal.region.id if assigned_terminal.region else None
        }]
        available_regions = [{'id': assigned_region.id, 'name': assigned_region.name}] if assigned_region else []

    
    terminals_for_frontend = [
        {
            'id': t['id'],
            'branch_name': t['branch_name'],
            'customer_name': t.get('customer__name') or getattr(t.get('customer'), 'name', 'N/A'),
            'region_name': t.get('region__name') or getattr(t.get('region'), 'name', 'N/A'),
            'region_id': t.get('region__id') or getattr(t.get('region'), 'id', None)
        }
        for t in available_terminals
    ]


    stats = {}

    stats["status_by_category"] = list(
        tickets.values("problem_category__name", "status")
        .annotate(count=Count("id"))
        .order_by("problem_category__name")
    )

   
    stats["problems_by_category"] = list(
        tickets.values("problem_category__name")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    stats["top_open_issues"] = list(
        tickets.filter(status="Open")
        .values("problem_category__name")
        .annotate(total=Count("id"))
        .order_by("-total")[:5]
    )

    # SLA Breaches
    sla_met = tickets.filter(
        is_sla_breached=False,
        status__iexact="closed"
    ).count()

    sla_breached_escalated = tickets.filter(
        is_sla_breached=True,
        is_escalated=True
    ).count()

    sla_breached_not_escalated = tickets.filter(
        is_sla_breached=True,
        is_escalated=False
    ).count()

    sla_met_before_escalation = tickets.filter(
        is_sla_breached=False,
        is_escalated=True,         
        status__iexact="closed"
    ).count()


    unresolved_tickets = tickets.filter(
        resolved_at__isnull=True, 
        due_date__isnull=False, 
        due_date__lt=timezone.now()
    )

    sla_escalated = tickets.filter(is_sla_breached=True, is_escalated=True).count()
    sla_not_escalated = tickets.filter(is_sla_breached=True, is_escalated=False).count()


    # Problem Categories
    categories = tickets.values("problem_category__name").annotate(
        count=Count("id")
    ).order_by("-count")

    # Created by (users)
    created_by_stats = tickets.values("created_by__username").annotate(
        count=Count("id")
    )

    # Tickets assigned to
    assignee_stats = (
        tickets.values("assigned_to__username")
        .annotate(count=Count("id"))
        .order_by("-count")
    )

    resolved_assignee_stats = (
        tickets.filter(status__iexact="Closed")  
        .values("assigned_to__username")
        .annotate(count=Count("id"))
        .order_by("-count")
    )

    resolved_by_stats = (
        tickets.filter(status__iexact="Closed")
        .values("resolved_by__username")
        .annotate(count=Count("id"))
        .order_by("-count")
    )

    unresolved_stats = tickets.filter(status__in=["Open", "In Progress"]) \
    .values("assigned_to__username") \
    .annotate(count=Count("id")) \
    .order_by("-count")

    # Replace None values with "Unassigned"
    unresolved_stats = [
        {**item, 'assigned_to__username': item.get('assigned_to__username') or 'Unassigned'}
        for item in unresolved_stats
    ]

    resolution_stats = tickets.aggregate(
        resolved=Count(Case(
            When(status__iexact="closed", then=1),
            output_field=IntegerField()
        )),
        unresolved=Count(Case(
            When(Q(status__iexact="open") | Q(status__iexact="in_progress"), then=1),
            output_field=IntegerField()
        ))
    )

   

    data_json = json.dumps(stats, cls=DjangoJSONEncoder)

    data = {
        'ticketsPerTerminal': [{'branch_name': entry['terminal__branch_name'], 'count': entry['ticket_count']} for entry in tickets_per_terminal],
        'ticketCategories': {'labels': [entry['problem_category__name'] for entry in ticket_categories], 'data': [entry['ticket_count'] for entry in ticket_categories]},
        "resolvedTickets": resolution_stats["resolved"] or 0,
        "unresolvedCount": resolution_stats["unresolved"] or 0,
        'ticketStatuses': {'labels': status_labels, 'data': status_counts},
        'days': [day.strftime('%Y-%m-%d') for day in days],
        'weekdays': weekdays,
        'hours': hours,
        'months': months,
        'years': years,
        "ticketsPerDay": {
            "labels": [day.strftime('%Y-%m-%d') for day in days],
            "data": tickets_per_day,
        },
        "ticketsPerWeekday": {
            "labels": weekdays,
            "data": tickets_per_weekday,
        },
        "ticketsPerHour": {
            "labels": hours,
            "data": tickets_per_hour,
        },
        "ticketsPerMonth": {
            "labels": months,
            "data": tickets_per_month,
        },
        "ticketsPerYear": {
            "labels": years,
            "data": tickets_per_year,
        },
        'terminals': terminals_for_frontend,
        'customers': available_customers,
        'regions': available_regions,
        'data_json': data_json,
         "escalationSLAStats": {  
            "labels": ["Breached & Escalated",
               "Breached & Not Escalated",
               "Met (incl. closed before escalation)"],
            "data": [
                sla_breached_escalated,
                sla_breached_not_escalated,
                sla_met + sla_met_before_escalation
            ]
        },
        # Problem categories
        "ticketCategories": {
            "labels": [c["problem_category__name"] for c in categories],
            "data": [c["count"] for c in categories],
        },

        # Tickets by creator
        "ticketsByCreator": {
            "labels": [c["created_by__username"] for c in created_by_stats],
            "data": [c["count"] for c in created_by_stats],
        },
        "ticketsByAssignee": {
            "labels": [a.get("assigned_to__username") or "Unassigned" for a in assignee_stats],
            "data": [a["count"] for a in assignee_stats],
        },
        "ticketsByResolver": {
            "labels": [r.get("resolved_by__username") or "Unresolved" for r in resolved_by_stats],
            "data": [r["count"] for r in resolved_by_stats],
        },
        "unresolvedTickets": {
            "labels": [u.get("assigned_to__username") or "Unassigned" for u in unresolved_stats] or ["No Unresolved"],
            "data": [u["count"] for u in unresolved_stats] or [0],
        },
        "resolvedByAssignee": {
            "labels": [r.get("assigned_to__username") or "Unassigned" for r in resolved_assignee_stats],
            "data": [r["count"] for r in resolved_assignee_stats],
        },
        "resolvedByResolver": {
            "labels": [r.get("resolved_by__username") or "Unresolved" for r in resolved_by_stats],
            "data": [r["count"] for r in resolved_by_stats],
        },

    }

    for ticket in tickets:
        print(f"Ticket {ticket.id}: resolved_at={ticket.resolved_at}, due_date={ticket.due_date}, is_sla_breached={ticket.is_sla_breached}, is_escalated={ticket.is_escalated}")



    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse(data, safe=False)

    """user_group = None
    if Customer.objects.filter(custodian=request.user).exists():
        user_group = "Custodian"
    elif Customer.objects.filter(overseer=request.user).exists():
        user_group = "Overseer"
    else:
        if request.user.groups.filter(name="Director").exists():
            user_group = "Director"
        elif request.user.groups.filter(name="Manager").exists():
            user_group = "Manager"
        elif request.user.groups.filter(name="Staff").exists():
            user_group = "Staff"
        else:
            user_group = "Customer"

    allowed_roles = ["Director", "Manager", "Staff", "Superuser"]"""

    return render(request, 'core/helpdesk/statistics.html', {
        "user_group": user_group,
        'customers': available_customers,
        'terminals': terminals_for_frontend,
        'regions': available_regions,
        "time_period": time_period,
        'selected_customer': str(customer_filter),
        'selected_terminal': str(terminal_filter),
        'selected_region': str(region_filter),
        'data_json': json.dumps(data, ensure_ascii=False),
        "assigned_customer": assigned_customer,
        "assigned_branch": assigned_terminal,
        "assigned_region": assigned_region,
        "hide_sla": hide_sla,
        "hide_assignee": hide_assignee,
        "hide_resolver": hide_resolver,
        "hide_terminal": hide_terminal
        #"allowed_roles": allowed_roles
    })

@login_required(login_url='login')
def export_report(request):
    import openpyxl
    from openpyxl.styles import Font, Border, Side
    from django.utils import timezone
    from django.db.models import Count, Case, When, IntegerField, Q
    from datetime import timedelta
    from django.http import HttpResponse

    today = timezone.now()
    user = request.user
    user_profile = getattr(user, "profile", None)

    # --- Role-based Filtering ---
    tickets = Ticket.objects.all()
    user_group = None
    assigned_customer = assigned_terminal = assigned_region = None

    if user.is_superuser or user.groups.filter(name__in=["Director", "Manager", "Staff"]).exists():
        user_group = "Internal"
    elif Customer.objects.filter(overseer=user).exists():
        user_group = "Overseer"
        assigned_customer = Customer.objects.filter(overseer=user).first()
        tickets = tickets.filter(terminal__customer=assigned_customer)
    elif user_profile and user_profile.terminal and user_profile.terminal.custodian == user:
        user_group = "Custodian"
        assigned_terminal = user_profile.terminal
        assigned_customer = assigned_terminal.customer
        assigned_region = assigned_terminal.region
        tickets = tickets.filter(terminal=assigned_terminal)
    else:
        tickets = Ticket.objects.none()

    # --- Filters ---
    tp = request.GET.get("time-period", "all_time")
    cust = request.GET.get("customer", "all")
    term = request.GET.get("terminal", "all")
    reg = request.GET.get("region", "all")

    # Date filter
    start, end = None, None
    if tp == "today":
        start = today.replace(hour=0, minute=0, second=0, microsecond=0)
        end = today.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif tp == "yesterday":
        start = (today - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        end = (today - timedelta(days=1)).replace(hour=23, minute=59, second=59, microsecond=999999)
    elif tp == "lastweek":
        start = today - timedelta(days=today.weekday() + 7)
        start = start.replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + timedelta(days=6, hours=23, minutes=59, seconds=59, microseconds=999999)
    elif tp == "lastmonth":
        end = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0) - timedelta(microseconds=1)
        start = end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif tp == "lastyear":
        start = today.replace(year=today.year - 1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        end = today.replace(year=today.year - 1, month=12, day=31, hour=23, minute=59, second=59, microsecond=999999)

    if start and end:
        tickets = tickets.filter(created_at__range=[start, end])

    if cust not in ["all", "", None]:
        tickets = tickets.filter(terminal__customer__id=cust)
    if term not in ["all", "", None]:
        tickets = tickets.filter(terminal__id=term)
    if reg not in ["all", "", None]:
        tickets = tickets.filter(terminal__region__id=reg)

    if not tickets.exists():
        return HttpResponse("No tickets to export matching your criteria.", status=404)

    # ========== Build all chart datasets ==========

    # Tickets by status
    status_qs = tickets.values("status").annotate(c=Count("id"))
    status_labels = [x["status"] for x in status_qs]
    status_counts = [x["c"] for x in status_qs]

    # Tickets per day / hour / month / year
    days = [today - timedelta(days=i) for i in range(7)]
    t_per_day = [tickets.filter(created_at__date=d.date()).count() for d in days]

    hours = [f"{h}-{h+1}" for h in range(24)]
    t_per_hour = [tickets.filter(created_at__hour=h).count() for h in range(24)]

    months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    t_per_month = [tickets.filter(created_at__month=i+1).count() for i in range(12)]

    years = sorted(set(t.created_at.year for t in tickets))
    t_per_year = [tickets.filter(created_at__year=y).count() for y in years]

    # Per terminal / category
    term_qs = tickets.values("terminal__branch_name").annotate(c=Count("id"))
    term_labels = [x["terminal__branch_name"] for x in term_qs]
    term_counts = [x["c"] for x in term_qs]

    cat_qs = tickets.values("problem_category__name").annotate(c=Count("id"))
    cat_labels = [x["problem_category__name"] for x in cat_qs]
    cat_counts = [x["c"] for x in cat_qs]

    # By creator / assignee / resolver
    creator_qs = tickets.values("created_by__username").annotate(c=Count("id"))
    assignee_qs = tickets.values("assigned_to__username").annotate(c=Count("id"))
    resolver_qs = tickets.filter(status__iexact="Closed").values("resolved_by__username").annotate(c=Count("id"))

    # Unresolved tickets by assignee
    unresolved_qs = tickets.filter(status__in=["Open","In Progress"]) \
                           .values("assigned_to__username") \
                           .annotate(c=Count("id")).order_by("-c")
    unresolved_labels = [u["assigned_to__username"] or "Unassigned" for u in unresolved_qs]
    unresolved_counts = [u["c"] for u in unresolved_qs]

    # SLA breach data
    sla_breached_escalated = tickets.filter(is_sla_breached=True, is_escalated=True).count()
    sla_breached_not = tickets.filter(is_sla_breached=True, is_escalated=False).count()
    sla_met = tickets.filter(is_sla_breached=False, status__iexact="closed").count()
    sla_labels = ["Breached & Escalated", "Breached & Not Escalated", "Met/Closed"]
    sla_counts = [sla_breached_escalated, sla_breached_not, sla_met]

    # ========== Write to Excel ==========

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Statistics"

    bold = Font(bold=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    def write(title, labels, values, row):
        ws.cell(row=row, column=1, value=title).font = bold
        for i, lbl in enumerate(labels):
            ws.cell(row=row + 1 + i, column=1, value=lbl)
            ws.cell(row=row + 1 + i, column=2, value=values[i])
        for r in range(row, row + 1 + len(labels)):
            for c in (1, 2):
                ws.cell(row=r, column=c).border = border
        return row + 2 + len(labels)

    r = 1
    r = write("Tickets by Status", status_labels, status_counts, r)
    r = write("Tickets per Terminal", term_labels, term_counts, r)
    r = write("Tickets per Category", cat_labels, cat_counts, r)
    r = write("Tickets per Day (7d)", [d.strftime("%Y-%m-%d") for d in days], t_per_day, r)
    r = write("Tickets per Hour", hours, t_per_hour, r)
    r = write("Tickets per Month", months, t_per_month, r)
    r = write("Tickets per Year", years, t_per_year, r)
    r = write("Tickets by Creator", [x["created_by__username"] for x in creator_qs],
              [x["c"] for x in creator_qs], r)
    r = write("Tickets by Assignee", [x["assigned_to__username"] or "Unassigned" for x in assignee_qs],
              [x["c"] for x in assignee_qs], r)
    r = write("Tickets by Resolver", [x["resolved_by__username"] or "Unresolved" for x in resolver_qs],
              [x["c"] for x in resolver_qs], r)
    r = write("Unresolved Tickets by Assignee", unresolved_labels, unresolved_counts, r)
    r = write("SLA Breach Stats", sla_labels, sla_counts, r)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="ticket_statistics.xlsx"'
    wb.save(response)
    return response